from flask import Flask, render_template, request, jsonify, send_file
import os
from collections import defaultdict
import datetime
import json
import pandas as pd
from werkzeug.utils import secure_filename
import chardet
import folium
import geopy
from geopy.geocoders import Nominatim
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import psycopg2
import psycopg2.extras
import io
import base64
import tempfile
import shutil
import zipfile

app = Flask(__name__)
geolocator = Nominatim(user_agent="bounding_box_app")
#Excel Reporting
UPLOAD_FOLDER = 'uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

def run_excel_qa(file_path):
    report_lines = []
    
    # File info
    file_name = os.path.basename(file_path)
    file_ext = os.path.splitext(file_path)[1]
    file_size_mb = round(os.path.getsize(file_path) / (1024 * 1024), 2)
    last_modified = datetime.fromtimestamp(os.path.getmtime(file_path)).strftime('%Y-%m-%d %H:%M:%S')

    report_lines.append(f"File Name: {file_name}")
    report_lines.append(f"File Extension: {file_ext}")
    report_lines.append(f"File Size: {file_size_mb} MB")
    report_lines.append(f"Last Modified: {last_modified}\n")

    wb = load_workbook(file_path, data_only=True)

    for sheet in wb.worksheets:
        sheet_name = sheet.title
        is_hidden = sheet.sheet_state != 'visible'
        display_name = f"{sheet_name} (hidden)" if is_hidden else sheet_name
        report_lines.append(f"Sheet: {display_name}")

        df = pd.DataFrame(sheet.values)
        num_rows, num_cols = df.shape
        report_lines.append(f"  - Total Rows: {num_rows}")
        report_lines.append(f"  - Total Columns: {num_cols}")

        # Empty cells
        empty_cells = []
        for r in range(1, num_rows + 1):
            for c in range(1, num_cols + 1):
                value = sheet.cell(row=r, column=c).value
                if value in [None, "", " "]:
                    cell_name = f"{get_column_letter(c)}{r}"
                    empty_cells.append(cell_name)
        report_lines.append(f"  - Empty Cells: {', '.join(empty_cells) if empty_cells else 'None'}")

        # Hidden columns
        hidden_columns = []
        for col_cells in sheet.iter_cols():
            col_letter = get_column_letter(col_cells[0].column)
            if sheet.column_dimensions[col_letter].hidden:
                hidden_columns.append(col_letter)
        report_lines.append(f"  - Hidden Columns: {', '.join(hidden_columns) if hidden_columns else 'None'}")

        # Data quality
        quality_issues = []
        for col in df.columns:
            col_series = df[col].dropna()
            if col_series.empty:
                quality_issues.append(f"    - Column {get_column_letter(col + 1)} is completely empty")
            elif col_series.nunique() == 1 and col_series.iloc[0] in [None, "", " "]:
                quality_issues.append(f"    - Column {get_column_letter(col + 1)} has uniform missing value")
            elif col_series.apply(type).nunique() > 1:
                quality_issues.append(f"    - Column {get_column_letter(col + 1)} has mixed data types")
        if quality_issues:
            report_lines.append("  - Data Quality Issues:")
            report_lines.extend(quality_issues)
        else:
            report_lines.append("  - Data Quality Issues: None")
        report_lines.append("")

    return "\n".join(report_lines)

#Functions list
# Report generator
def generate_folder_report(folder_path):
    file_count = 0
    total_size = 0
    file_types = defaultdict(list)
    empty_folders = []
    corrupt_files = []
    hidden_folders = []

    folder_name = os.path.basename(folder_path)
    folder_creation_time = os.path.getctime(folder_path)
    folder_last_modified_time = os.path.getmtime(folder_path)

    for dirpath, dirnames, filenames in os.walk(folder_path):
        for dirname in dirnames:
            if dirname.startswith('.'):
                hidden_folders.append(dirname)

        if not filenames and not dirnames:
            empty_folders.append(dirpath)

        for filename in filenames:
            file_count += 1
            _, ext = os.path.splitext(filename)
            file_types[ext].append(f"{filename} ({dirpath})")

            file_path = os.path.join(dirpath, filename)
            total_size += os.path.getsize(file_path)

            try:
                with open(file_path, 'rb') as f:
                    f.read(1)
            except Exception:
                corrupt_files.append(file_path)

    size_kb = total_size / 1024
    size_mb = size_kb / 1024
    size_gb = size_mb / 1024

    report = {
        "Folder Name": folder_name,
        "Hidden_Folders": {"Count": len(hidden_folders), "Names": hidden_folders},
        "Creation Date": datetime.datetime.fromtimestamp(folder_creation_time).strftime('%Y-%m-%d %H:%M:%S'),
        "Last Modified Date": datetime.datetime.fromtimestamp(folder_last_modified_time).strftime('%Y-%m-%d %H:%M:%S'),
        "Total Files": file_count,
        "Total Size (bytes)": total_size,
        "Total Size (KB)": size_kb,
        "Total Size (MB)": size_mb,
        "Total Size (GB)": size_gb,
        "File Types": dict(file_types),
        "Empty Folders": empty_folders,
        "Corrupt Files": corrupt_files,
        "Folder Structure Issues": []
    }

    if empty_folders or corrupt_files:
        report["Folder Structure Issues"].append("Issues found:")
        if empty_folders:
            report["Folder Structure Issues"].append(f"Empty folders: {len(empty_folders)}")
        if corrupt_files:
            report["Folder Structure Issues"].append(f"Corrupt files: {len(corrupt_files)}")

    return report
# Function to convert JSON to CSV
# Function to convert JSON file object to a CSV and return path
def json_to_csv(file):
    data = json.load(file)
    df = pd.json_normalize(data)

    # Create a temporary file to store CSV
    temp_csv = tempfile.NamedTemporaryFile(delete=False, suffix=".csv", mode='w', newline='', encoding='utf-8')
    df.to_csv(temp_csv.name, index=False)
    return temp_csv.name


# Function to detect file encoding
def detect_encoding(file_path):
    with open(file_path, 'rb') as file:
        raw_data = file.read()
        result = chardet.detect(raw_data)
        encoding = result['encoding']
        confidence = result['confidence']
        return encoding, confidence

#Function to add a geographic bounding box  
def create_bounding_box(lat, lon, delta=0.1):
    """Create a bounding box around a point."""
    return {
        "west": lon - delta,
        "east": lon + delta,
        "south": lat - delta,
        "north": lat + delta
    }

def generate_bounding_box(points):
    """Create a bounding box around multiple points."""
    lats = [point[0] for point in points]
    lons = [point[1] for point in points]
    return {
        "west": min(lons),
        "east": max(lons),
        "south": min(lats),
        "north": max(lats)
    }


@app.route('/', methods=['GET', 'POST'])
def index():
    report = {}
    if request.method == 'POST':
        folder_path = request.form['folder_path']
        report = generate_folder_report(folder_path)
    return render_template('index.html', folder_report=report)

@app.route('/count', methods=['POST'])
def count():
    text = request.form.get('text', '')
    word_count = len(text.split()) if text else 0
    char_count = len(text)
    return jsonify({'words': word_count, 'characters': char_count})

# Endpoint to receive ZIP file and return folder report
@app.route('/folder_report', methods=['POST'])
def folder_report():
    if 'folder_zip' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400

    zip_file = request.files['folder_zip']
    if zip_file.filename == '' or not zip_file.filename.endswith('.zip'):
        return jsonify({'error': 'Invalid file type. Only ZIP files allowed.'}), 400

    # Extract ZIP to temp directory
    temp_dir = tempfile.mkdtemp()
    try:
        zip_path = os.path.join(temp_dir, zip_file.filename)
        zip_file.save(zip_path)

        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            zip_ref.extractall(temp_dir)

        # Find root folder in ZIP
        extracted_folders = [f.path for f in os.scandir(temp_dir) if f.is_dir()]
        if not extracted_folders:
            return jsonify({'error': 'No folders found in ZIP file.'}), 400

        report = generate_folder_report(extracted_folders[0])
        return jsonify(report)

    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        shutil.rmtree(temp_dir)


@app.route('/generate', methods=['GET', 'POST'])
def generate():
    if request.method == 'POST':
        folder_path = request.form['folder_path']
        if 'generate' in request.form:
            alart_message = paths_generator(folder_path)
            return render_template('index.html', alart_message=alart_message)

@app.route('/geobox', methods=['POST'])
def geobox():
    location_input = request.json['coordinates']
    location = geolocator.geocode(location_input)

    if not location:
        return jsonify({'error': 'Location not found'}), 404

    lat, lon = location.latitude, location.longitude
    bounding_box = create_bounding_box(lat, lon)

    # Create a map with the bounding box
    m = folium.Map(location=[lat, lon], zoom_start=10)
    folium.Rectangle(
        bounds=[[bounding_box['south'], bounding_box['west']],
                [bounding_box['north'], bounding_box['east']]],
        color='blue',
        fill=True,
        fill_opacity=0.2
    ).add_to(m)

    map_html = m._repr_html_()

    return jsonify({
        'bounding_box': bounding_box,
        'map_html': map_html
    })

@app.route('/geobox_multiple', methods=['POST'])
def geobox_multiple():
    data = request.get_json()
    coordinates_input = data.get('coordinates')  # Expecting a string like "lat1,lon1;lat2,lon2;..."

    try:
        coordinates_list = [tuple(map(float, coord.split(','))) for coord in coordinates_input.split(';')]
    except Exception:
        return jsonify({'error': 'Invalid coordinates format'}), 400

    if not coordinates_list:
        return jsonify({'error': 'No coordinates provided'}), 400

    bounding_box_multiple = generate_bounding_box(coordinates_list)

    # Create a map with the bounding box
    m = folium.Map(
        location=[(bounding_box_multiple['north'] + bounding_box_multiple['south']) / 2, 
                  (bounding_box_multiple['east'] + bounding_box_multiple['west']) / 2],
        zoom_start=10
    )
    folium.Rectangle(
        bounds=[[bounding_box_multiple['south'], bounding_box_multiple['west']],
                [bounding_box_multiple['north'], bounding_box_multiple['east']]],
        color='blue',
        fill=True,
        fill_opacity=0.2
    ).add_to(m)

    map_html_multiple = m._repr_html_()

    return jsonify({
        'bounding_box': bounding_box_multiple,
        'map_html': map_html_multiple
    })

@app.route("/Excel_reporting", methods=["GET", "POST"])
def Excel_reporter():
    report = None
    if request.method == "POST":
        uploaded_file = request.files["excel_file"]
        if uploaded_file and uploaded_file.filename.endswith((".xlsx", ".xls")):
            filename = secure_filename(uploaded_file.filename)
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            uploaded_file.save(file_path)
            report = run_excel_qa(file_path)
    return render_template("index.html", excel_report=report)

@app.route('/leave_balance', methods=['GET', 'POST'])
def leave_balance():
    result = None
    if request.method == 'POST':
        try:
            total_hours = float(request.form['total_hours'])
            daily_hours = float(request.form['daily_hours'])
            days_left = total_hours / daily_hours
            result = round(days_left, 2)
        except (ValueError, ZeroDivisionError):
            result = "Invalid input. Please enter valid numbers."

    return render_template('index.html', leave=result)

def get_db_connection():
    return psycopg2.connect(
        dbname='AGT',
        user='postgres',
        password='pgsqtk116chuk95',
        host='chukspace.ctiuisa62ks5.eu-north-1.rds.amazonaws.com',
        port='5432'
    )

# Fetch data for dashboard
def get_sample_data():
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    # Tasks
    cur.execute('SELECT title, status FROM "CDMA".tasks ORDER BY id')
    tasks = [{'title': row['title'], 'status': row['status']} for row in cur.fetchall()]

    # Issues
    cur.execute('SELECT id, description, status FROM "CDMA".issues ORDER BY id')
    issues = [{'id': row['id'], 'description': row['description'], 'status': row['status']} for row in cur.fetchall()]

    # Events
    cur.execute('SELECT title, start FROM "CDMA".calendar_events ORDER BY start')
    events = [{'title': row['title'], 'start': row['start'].isoformat()} for row in cur.fetchall()]

    cur.close()
    conn.close()
    return tasks, issues, events

# Dashboard route
@app.route('/manager')
def manager():
    tasks, issues, events = get_sample_data()
    return render_template('index.html', tasks=tasks, issues=issues, events=events)

# -----------------------------
# EVENT ROUTES
# -----------------------------
@app.route('/events', methods=['GET'])
def get_events():
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute('SELECT title, start FROM "CDMA".calendar_events ORDER BY start')
    events = [{'title': row['title'], 'start': row['start'].isoformat()} for row in cur.fetchall()]
    cur.close()
    conn.close()
    return jsonify(events)

@app.route('/add_event', methods=['POST'])
def add_event():
    data = request.get_json()
    title = data.get('title')
    start = data.get('start')
    if not title or not start:
        return jsonify({'status': 'error', 'message': 'Missing data'}), 400

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('INSERT INTO "CDMA".calendar_events (title, start) VALUES (%s, %s)', (title, start))
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({'status': 'success', 'event': {'title': title, 'start': start}}), 201

# -----------------------------
# TASK ROUTES
# -----------------------------
@app.route('/add_task', methods=['POST'])
def add_task():
    data = request.get_json()
    title = data.get('title')
    status = data.get('status')
    if not title or not status:
        return jsonify({'status': 'error', 'message': 'Missing task data'}), 400

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('INSERT INTO "CDMA".tasks (title, status) VALUES (%s, %s)', (title, status))
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({'status': 'success', 'task': {'title': title, 'status': status}}), 201

# -----------------------------
# ISSUE ROUTES
# -----------------------------
@app.route('/add_issue', methods=['POST'])
def add_issue():
    data = request.get_json()
    description = data.get('description')
    status = data.get('status')
    if not description or not status:
        return jsonify({'status': 'error', 'message': 'Missing issue data'}), 400

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        'INSERT INTO "CDMA".issues (description, status) VALUES (%s, %s) RETURNING id',
        (description, status)
    )
    issue_id = cur.fetchone()[0]
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({
        'status': 'success',
        'issue': {'id': issue_id, 'description': description, 'status': status}
    }), 201
    
@app.route('/delete_task/<int:task_id>', methods=['DELETE'])
def delete_task(task_id):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('DELETE FROM "CDMA".tasks WHERE id = %s', (task_id,))
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({'status': 'success'})

# NEW: Hide/Delete Issue
@app.route('/delete_issue/<int:issue_id>', methods=['DELETE'])
def delete_issue(issue_id):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('DELETE FROM "CDMA".issues WHERE id = %s', (issue_id,))
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({'status': 'success'})

# NEW: Add Achievement
@app.route('/add_achievement', methods=['POST'])
def add_achievement():
    data = request.get_json()
    required_fields = ['task', 'description', 'project', 'manager', 'date_started', 'date_ended', 'status', 'comment']
    if not all(data.get(f) for f in required_fields):
        return jsonify({'status': 'error', 'message': 'Missing achievement data'}), 400

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('''
        INSERT INTO "CDMA".achievements 
        (task, description, project, project_manager, date_started, date_ended, completion_status, comment)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
    ''', (
        data['task'], data['description'], data['project'], data['manager'],
        data['date_started'], data['date_ended'], data['status'], data['comment']
    ))
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({'status': 'success'})

if __name__ == '__main__':
    app.run(debug=True)

